import openpyxl

class Excel():

    def get_total_row(self,filepath,sheetname):
        wb = openpyxl.open(filepath)
        sheet = wb[sheetname]
        return sheet.max_row

    def get_total_col(self,filepath,sheetname):
        wb = openpyxl.open(filepath)
        sheet = wb[sheetname]
        return sheet.max_column

    def get_value(self,filepath,sheetname,row,col):
        wb = openpyxl.open(filepath)
        sheet = wb[sheetname]
        return sheet.cell(row,col).value

    def set_value(self,filepath,sheetname,row,col,data):
        wb = openpyxl.open(filepath)
        sheet = wb[sheetname]
        sheet.cell(row,col).value = data
        wb.save(filepath)
        wb.close()